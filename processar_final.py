# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

def processar_planilha():
    try:
        print("INICIANDO PROCESSAMENTO...")
        
        # Carregar arquivo
        wb = load_workbook('planilhafluxo.xlsx', data_only=True)
        
        if 'fluxortd' not in wb.sheetnames:
            print("ABA fluxortd NAO ENCONTRADA")
            return False
            
        ws = wb['fluxortd']
        print("ABA fluxortd ENCONTRADA")
        
        # Coletar dados CHEIO (colunas 1-7)
        dados_cheio = []
        for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
            if row and any(row):
                dados_cheio.append(row)
        
        print(f"DADOS CHEIO: {len(dados_cheio)} linhas")
        
        # Coletar dados MINI (colunas 9-15) 
        dados_mini = []
        for row in ws.iter_rows(min_row=3, min_col=9, max_col=15, values_only=True):
            if row and any(row):
                dados_mini.append(row)
        
        print(f"DADOS MINI: {len(dados_mini)} linhas")
        
        # Atualizar aba saldo se existir
        if 'saldo' in wb.sheetnames:
            print("ATUALIZANDO ABA SALDO...")
            # Aqui viria a logica de atualizacao
            pass
        
        # Salvar resultado
        wb.save('planilhafluxo_PROCESSADO.xlsx')
        print("ARQUIVO SALVO: planilhafluxo_PROCESSADO.xlsx")
        return True
        
    except Exception as e:
        print(f"ERRO: {e}")
        return False

# Executar
if __name__ == "__main__":
    success = processar_planilha()
    if success:
        print("PROCESSAMENTO CONCLUIDO COM SUCESSO!")
    else:
        print("FALHA NO PROCESSAMENTO")