# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

def calcular_saldos_contrato(dados, nome_contrato):
    """Calcula saldos para um contrato específico"""
    if not dados:
        print(f"SEM DADOS PARA {nome_contrato}")
        return pd.DataFrame()
    
    # Criar DataFrame
    colunas = ['Data', 'Compradora', 'Valor', 'Quantidade', 'Vendedora', 'Agressor', 'Agente']
    df = pd.DataFrame(dados, columns=colunas)
    
    # Converter para numérico
    df['Valor'] = pd.to_numeric(df['Valor'], errors='coerce')
    df['Quantidade'] = pd.to_numeric(df['Quantidade'], errors='coerce')
    df = df.dropna(subset=['Valor', 'Quantidade'])
    df = df[(df['Valor'] > 0) & (df['Quantidade'] > 0)]
    
    if len(df) == 0:
        print(f"DADOS {nome_contrato} INVALIDOS APOS LIMPEZA")
        return pd.DataFrame()
    
    # Calcular fluxo financeiro
    df['Financeiro'] = df['Valor'] * df['Quantidade']
    
    # Agrupar por corretora
    compras = df.groupby('Compradora')['Financeiro'].sum()
    vendas = df.groupby('Vendedora')['Financeiro'].sum()
    
    # Calcular saldos
    todas_corretoas = set(compras.index).union(set(vendas.index))
    resultados = []
    
    for corretora in todas_corretoas:
        compra = compras.get(corretora, 0)
        venda = vendas.get(corretora, 0)
        saldo = compra - venda
        
        # Calcular volume
        vol_compra = df[df['Compradora'] == corretora]['Quantidade'].sum()
        vol_venda = df[df['Vendedora'] == corretora]['Quantidade'].sum()
        vol_liquido = vol_compra - vol_venda
        
        # Preço médio
        preco_medio = saldo / vol_liquido if vol_liquido != 0 else 0
        
        resultados.append({
            'Corretora': corretora,
            'Vol. Qtd': vol_liquido,
            'Média': preco_medio,
            'Financeiro': saldo
        })
    
    df_resultado = pd.DataFrame(resultados)
    print(f"{nome_contrato} - {len(df_resultado)} CORRETORAS - SALDO: R$ {df_resultado['Financeiro'].sum():,.2f}")
    
    return df_resultado

def processar_planilha_completa():
    try:
        print("INICIANDO PROCESSAMENTO COMPLETO...")
        
        # Carregar arquivo
        wb = load_workbook('planilhafluxo.xlsx', data_only=True)
        
        if 'fluxortd' not in wb.sheetnames:
            print("ABA fluxortd NAO ENCONTRADA")
            return False
            
        ws = wb['fluxortd']
        print("ABA fluxortd ENCONTRADA")
        
        # Coletar dados CHEIO (colunas 1-7)
        dados_cheio = []
        for row in ws.iter_rows(min_row=3, max_col=7, values_only=True):
            if row and any(row):
                dados_cheio.append(row)
        
        # Coletar dados MINI (colunas 9-15)
        dados_mini = []
        for row in ws.iter_rows(min_row=3, min_col=9, max_col=15, values_only=True):
            if row and any(row):
                dados_mini.append(row)
        
        # Processar ambos os contratos
        resultados_cheio = calcular_saldos_contrato(dados_cheio, "CHEIO")
        resultados_mini = calcular_saldos_contrato(dados_mini, "MINI")
        
        # Atualizar aba saldo
        if 'saldo' in wb.sheetnames:
            ws_saldo = wb['saldo']
            print("ATUALIZANDO ABA SALDO...")
            
            # Limpar dados antigos
            for row in range(4, 100):
                try:
                    # Limpar área CHEIO (colunas A-F)
                    if ws_saldo[f'A{row}'].value not in ['Estrangeiros', 'Bancos', 'Pessoas Físicas', 'Saldo']:
                        ws_saldo[f'A{row}'].value = None
                        ws_saldo[f'B{row}'].value = None
                        ws_saldo[f'C{row}'].value = None
                        ws_saldo[f'E{row}'].value = None
                    
                    # Limpar área MINI (colunas G-L)
                    if ws_saldo[f'G{row}'].value not in ['Estrangeiros', 'Bancos', 'Pessoas Físicas', 'Saldo']:
                        ws_saldo[f'G{row}'].value = None
                        ws_saldo[f'H{row}'].value = None
                        ws_saldo[f'I{row}'].value = None
                        ws_saldo[f'K{row}'].value = None
                except:
                    continue
            
            # Preencher CHEIO
            if len(resultados_cheio) > 0:
                linha = 4
                for _, row in resultados_cheio.iterrows():
                    if linha <= 100:
                        ws_saldo[f'A{linha}'] = row['Corretora']
                        ws_saldo[f'B{linha}'] = row['Vol. Qtd']
                        ws_saldo[f'C{linha}'] = row['Média']
                        ws_saldo[f'E{linha}'] = row['Financeiro']
                        linha += 1
            
            # Preencher MINI
            if len(resultados_mini) > 0:
                linha = 4
                for _, row in resultados_mini.iterrows():
                    if linha <= 100:
                        ws_saldo[f'G{linha}'] = row['Corretora']
                        ws_saldo[f'H{linha}'] = row['Vol. Qtd']
                        ws_saldo[f'I{linha}'] = row['Média']
                        ws_saldo[f'K{linha}'] = row['Financeiro']
                        linha += 1
        
        # Salvar arquivo
        wb.save('planilhafluxo_PROCESSADO_COMPLETO.xlsx')
        print("ARQUIVO SALVO: planilhafluxo_PROCESSADO_COMPLETO.xlsx")
        
        # Mostrar resultados
        if len(resultados_cheio) > 0:
            print("\nTOP 5 CHEIO:")
            top5 = resultados_cheio.nlargest(5, 'Financeiro')
            for i, (_, row) in enumerate(top5.iterrows(), 1):
                sinal = "+" if row['Financeiro'] > 0 else ""
                print(f"{i}. {row['Corretora']:15} {sinal}R$ {abs(row['Financeiro']):,}")
        
        if len(resultados_mini) > 0:
            print("\nTOP 5 MINI:")
            top5 = resultados_mini.nlargest(5, 'Financeiro')
            for i, (_, row) in enumerate(top5.iterrows(), 1):
                sinal = "+" if row['Financeiro'] > 0 else ""
                print(f"{i}. {row['Corretora']:15} {sinal}R$ {abs(row['Financeiro']):,}")
        
        return True
        
    except Exception as e:
        print(f"ERRO: {e}")
        import traceback
        traceback.print_exc()
        return False

# Executar
if __name__ == "__main__":
    success = processar_planilha_completa()
    if success:
        print("\nPROCESSAMENTO CONCLUIDO COM SUCESSO!")
    else:
        print("\nFALHA NO PROCESSAMENTO")