import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

def processar_planilha_fluxo(nome_arquivo):
    try:
        print(f"Carregando arquivo: {nome_arquivo}")
        
        wb = load_workbook(nome_arquivo, data_only=True)
        
        if 'fluxortd' not in wb.sheetnames:
            print("❌ Aba 'fluxortd' não encontrada")
            return None, None
            
        ws_fluxo = wb['fluxortd']
        print("✅ Estrutura detectada: 15 colunas (CHEIO + MINI)")
        
        # Coletar dados
        dados = []
        for row in ws_fluxo.iter_rows(min_row=3, values_only=True):
            if row and any(row):
                dados.append(row)
        
        # Criar DataFrame com todas as colunas
        colunas = ['Data_CHEIO', 'Compradora_CHEIO', 'Valor_CHEIO', 'Quantidade_CHEIO', 
                  'Vendedora_CHEIO', 'Agressor_CHEIO', 'Agente_CHEIO', 'Separador',
                  'Data_MINI', 'Compradora_MINI', 'Valor_MINI', 'Quantidade_MINI',
                  'Vendedora_MINI', 'Agressor_MINI', 'Agente_MINI']
        
        df = pd.DataFrame(dados, columns=colunas)
        
        # PROCESSAR CONTRATO CHEIO (colunas 1-7)
        print("\n📊 PROCESSANDO CONTRATO CHEIO...")
        df_cheio = df[['Compradora_CHEIO', 'Vendedora_CHEIO', 'Valor_CHEIO', 'Quantidade_CHEIO']].copy()
        df_cheio = df_cheio.rename(columns={
            'Compradora_CHEIO': 'Compradora',
            'Vendedora_CHEIO': 'Vendedora', 
            'Valor_CHEIO': 'Valor',
            'Quantidade_CHEIO': 'Quantidade'
        })
        
        # Converter para numérico e limpar
        df_cheio['Valor'] = pd.to_numeric(df_cheio['Valor'], errors='coerce')
        df_cheio['Quantidade'] = pd.to_numeric(df_cheio['Quantidade'], errors='coerce')
        df_cheio = df_cheio.dropna(subset=['Valor', 'Quantidade'])
        df_cheio = df_cheio[(df_cheio['Valor'] > 0) & (df_cheio['Quantidade'] > 0)]
        
        print(f"📈 Dados CHEIO válidos: {len(df_cheio)} linhas")
        
        # PROCESSAR CONTRATO MINI (colunas 9-15)
        print("\n📊 PROCESSANDO CONTRATO MINI...")
        df_mini = df[['Compradora_MINI', 'Vendedora_MINI', 'Valor_MINI', 'Quantidade_MINI']].copy()
        df_mini = df_mini.rename(columns={
            'Compradora_MINI': 'Compradora',
            'Vendedora_MINI': 'Vendedora', 
            'Valor_MINI': 'Valor',
            'Quantidade_MINI': 'Quantidade'
        })
        
        # Converter para numérico e limpar
        df_mini['Valor'] = pd.to_numeric(df_mini['Valor'], errors='coerce')
        df_mini['Quantidade'] = pd.to_numeric(df_mini['Quantidade'], errors='coerce')
        df_mini = df_mini.dropna(subset=['Valor', 'Quantidade'])
        df_mini = df_mini[(df_mini['Valor'] > 0) & (df_mini['Quantidade'] > 0)]
        
        print(f"📈 Dados MINI válidos: {len(df_mini)} linhas")
        
        # FUNÇÃO PARA CALCULAR SALDOS
        def calcular_saldos(df_tipo, nome_tipo):
            if len(df_tipo) == 0:
                print(f"⚠️  Nenhum dado válido para {nome_tipo}")
                return pd.DataFrame()
            
            # Calcular fluxo
            compras = df_tipo.groupby('Compradora').apply(
                lambda x: (x['Valor'] * x['Quantidade']).sum()
            ).reset_index(name='Financeiro_Compra')
            
            vendas = df_tipo.groupby('Vendedora').apply(
                lambda x: (x['Valor'] * x['Quantidade']).sum()
            ).reset_index(name='Financeiro_Venda')
            
            # Juntar e calcular saldo LÍQUIDO
            todas_corretoas = set(compras['Compradora']).union(set(vendas['Vendedora']))
            
            dados_corretoas = []
            for corretora in todas_corretoas:
                compra_valor = compras[compras['Compradora'] == corretora]['Financeiro_Compra'].sum()
                venda_valor = vendas[vendas['Vendedora'] == corretora]['Financeiro_Venda'].sum()
                saldo_liquido = compra_valor - venda_valor
                
                # Calcular volume líquido
                vol_compra = df_tipo[df_tipo['Compradora'] == corretora]['Quantidade'].sum()
                vol_venda = df_tipo[df_tipo['Vendedora'] == corretora]['Quantidade'].sum()
                vol_liquido = vol_compra - vol_venda
                
                # Preço médio
                preco_medio = saldo_liquido / vol_liquido if vol_liquido != 0 else 0
                
                dados_corretoas.append({
                    'Corretora': corretora,
                    'Vol. Qtd': vol_liquido,
                    'Média': preco_medio,
                    'Financeiro': saldo_liquido
                })
            
            df_resultado = pd.DataFrame(dados_corretoas)
            print(f"💰 {nome_tipo} - Saldo total: R$ {df_resultado['Financeiro'].sum():,.2f}")
            
            return df_resultado
        
        # Calcular saldos para ambos os contratos
        df_cheio_resultado = calcular_saldos(df_cheio, "CHEIO")
        df_mini_resultado = calcular_saldos(df_mini, "MINI")
        
        # ATUALIZAR ABA SALDO
        if 'saldo' in wb.sheetnames and (len(df_cheio_resultado) > 0 or len(df_mini_resultado) > 0):
            ws_saldo = wb['saldo']
            
            print("\n📝 ATUALIZANDO ABA SALDO...")
            
            # Limpar dados antigos
            for row in range(4, 100):
                try:
                    # Limpar área CHEIO (colunas A-F)
                    if ws_saldo[f'A{row}'].value not in ['Estrangeiros', 'Bancos', 'Pessoas Físicas', 'Saldo']:
                        ws_saldo[f'A{row}'].value = None
                        ws_saldo[f'B{row}'].value = None
                        ws_saldo[f'C{row}'].value = None
                        ws_saldo[f'E{row}'].value = None
                    
                    # Limpar área MINI (colunas G-L)
                    if ws_saldo[f'G{row}'].value not in ['Estrangeiros', 'Bancos', 'Pessoas Físicas', 'Saldo']:
                        ws_saldo[f'G{row}'].value = None
                        ws_saldo[f'H{row}'].value = None
                        ws_saldo[f'I{row}'].value = None
                        ws_saldo[f'K{row}'].value = None
                except:
                    pass
            
            # Preencher CHEIO (colunas A-F)
            if len(df_cheio_resultado) > 0:
                linha = 4
                for _, row_data in df_cheio_resultado.iterrows():
                    if linha <= 100:
                        try:
                            ws_saldo[f'A{linha}'] = row_data['Corretora']
                            ws_saldo[f'B{linha}'] = row_data['Vol. Qtd']
                            ws_saldo[f'C{linha}'] = row_data['Média']
                            ws_saldo[f'E{linha}'] = row_data['Financeiro']
                            linha += 1
                        except:
                            linha += 1
            
            # Preencher MINI (colunas G-L)
            if len(df_mini_resultado) > 0:
                linha = 4
                for _, row_data in df_mini_resultado.iterrows():
                    if linha <= 100:
                        try:
                            ws_saldo[f'G{linha}'] = row_data['Corretora']
                            ws_saldo[f'H{linha}'] = row_data['Vol. Qtd']
                            ws_saldo[f'I{linha}'] = row_data['Média']
                            ws_saldo[f'K{linha}'] = row_data['Financeiro']
                            linha += 1
                        except:
                            linha += 1
            
            print("✅ Dados atualizados na aba 'saldo'")
        
        # Salvar arquivo
        novo_nome = nome_arquivo.replace('.xlsx', '_PROCESSADO.xlsx')
        wb.save(novo_nome)
        
        print(f"\n💾 Arquivo salvo: {novo_nome}")
        
        return df_cheio_resultado, df_mini_resultado
        
    except Exception as e:
        print(f"❌ Erro: {e}")
        import traceback
        traceback.print_exc()
        return None, None

# EXECUTAR
if __name__ == "__main__":
    print("🚀 PROCESSANDO CHEIO + MINI")
    print("===========================")
    
    resultado_cheio, resultado_mini = processar_planilha_fluxo("planilhafluxo.xlsx")
    
    if resultado_cheio is not None and resultado_mini is not None:
        print("\n🎉 PROCESSAMENTO CONCLUÍDO!")
        print("===========================")
        
        if len(resultado_cheio) > 0:
            print("\n🏆 TOP 5 - CHEIO:")
            top5_cheio = resultado_cheio.nlargest(5, 'Financeiro')
            for i, (_, row) in enumerate(top5_cheio.iterrows(), 1):
                sinal = "+" if row['Financeiro'] > 0 else ""
                print(f"{i:2d}. {row['Corretora']:15} {sinal}R$ {abs(row['Financeiro']):12,.2f}")
        
        if len(resultado_mini) > 0:
            print("\n🏆 TOP 5 - MINI:")
            top5_mini = resultado_mini.nlargest(5, 'Financeiro')
            for i, (_, row) in enumerate(top5_mini.iterrows(), 1):
                sinal = "+" if row['Financeiro'] > 0 else ""
                print(f"{i:2d}. {row['Corretora']:15} {sinal}R$ {abs(row['Financeiro']):12,.2f}")
    else:
        print("\n❌ Falha no processamento")