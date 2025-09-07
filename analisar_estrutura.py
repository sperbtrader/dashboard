import pandas as pd
from openpyxl import load_workbook

def analisar_estrutura(nome_arquivo):
    print("🔍 ANALISANDO ESTRUTURA DA PLANILHA...")
    
    try:
        wb = load_workbook(nome_arquivo, data_only=True)
        
        if 'fluxortd' not in wb.sheetnames:
            print("❌ Aba 'fluxortd' não encontrada")
            return None
            
        ws_fluxo = wb['fluxortd']
        
        # Verificar cabeçalhos reais (linha 2)
        print("📋 Cabeçalhos encontrados (Linha 2):")
        cabecalhos = []
        for cell in ws_fluxo[2]:  # Linha 2 (cabeçalhos)
            cabecalhos.append(cell.value)
        
        print(f"Total de colunas: {len(cabecalhos)}")
        for i, cabecalho in enumerate(cabecalhos, 1):
            print(f"  Coluna {i:2d}: {cabecalho}")
        
        # Verificar primeira linha de dados (linha 3)
        print("\n📊 Primeira linha de dados (Linha 3):")
        primeira_linha = []
        for cell in ws_fluxo[3]:  # Linha 3 (primeiros dados)
            primeira_linha.append(cell.value)
        
        for i, valor in enumerate(primeira_linha, 1):
            print(f"  Coluna {i:2d}: {valor}")
        
        # Verificar algumas linhas para ver padrão
        print("\n🔍 Amostra de dados (5 primeiras linhas):")
        for row_num in range(3, 8):  # Linhas 3 a 7
            linha = []
            for cell in ws_fluxo[row_num]:
                linha.append(cell.value)
            print(f"Linha {row_num}: {linha}")
        
        return len(cabecalhos)
        
    except Exception as e:
        print(f"❌ Erro: {e}")
        return None

# Executar análise
if __name__ == "__main__":
    print("🚀 ANALISADOR DE ESTRUTURA DA PLANILHA")
    print("=======================================")
    
    num_colunas = analisar_estrutura("planilhafluxo.xlsx")
    
    if num_colunas:
        print(f"\n🎯 A planilha tem {num_colunas} colunas!")
        print("\n💡 Com base na análise, vamos criar o script correto.")
    else:
        print("\n❌ Não foi possível analisar a planilha.")